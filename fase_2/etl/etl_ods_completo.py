"""
ETL completo: Datos fuentes → ODS (ods_justicia)
=================================================
Fuentes procesadas:
  1. fase_1/datos_fuentes/contratos/       (CSVs OCDS)
     → Tipo_Contrato, Institucion, Empresa (proveedores OCDS), Contrato
  2. fase_1/datos_fuentes/finanzas/
     registro_de_proveedores.xlsx
     → Perfil_Persona (personas individuales), Empresa (todos los proveedores)
  3. fase_1/datos_fuentes/ine/
     procesos_judiciales.xlsx
     → Indicador_Externo (estadísticas de detenidos y tasas de incidencia)

Salidas (en fase_2/output/):
  csv/  → un CSV por tabla
  sql/  → un archivo SQL por fuente con los INSERTs correspondientes

NOTAS:
  - Los IDs de Contrato.id_trabajador y Contrato.id_cargo se dejan en
    placeholder (1) porque OCDS no contiene información de nómina.
  - El rango de IDs está particionado por fuente para evitar colisiones:
      OCDS:         Institucion 1-9999 / Empresa  1-9999      / Contrato 1-9999
      Proveedores:  Perfil_Persona 1..N / Empresa 10000..N
  - CLASIFICACION del RGAE se mapea a Tipo_Empresa (catálogo insert_ods.sql):
      1=Sociedad Anónima, 2=Empresa Individual, 3=SRL,
      4=Cooperativa,     5=ONG,                6=Consorcio
"""

import csv
import os
import re
from datetime import datetime, date
import openpyxl

# ── Rutas base ─────────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.abspath(__file__))
FUENTES   = os.path.join(BASE, "..", "fase_1", "datos_fuentes")
DIR_CONT  = os.path.join(FUENTES, "contratos")
DIR_FIN   = os.path.join(FUENTES, "finanzas")
DIR_INE   = os.path.join(FUENTES, "ine")

OUT_CSV   = os.path.join(BASE, "output", "csv")
OUT_SQL   = os.path.join(BASE, "output", "sql")

os.makedirs(OUT_CSV, exist_ok=True)
os.makedirs(OUT_SQL, exist_ok=True)

# ── Helpers ────────────────────────────────────────────────────────────────────
def esc(value) -> str:
    """Escapa un valor para SQL: None → NULL, cadenas con comillas escapadas."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def parse_date(raw) -> str | None:
    """Normaliza fechas ISO 8601 o datetime a 'YYYY-MM-DD'. Retorna None si vacío."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return None
    return s[:10]


def extract_nit(raw_id: str) -> str:
    """'GT-NIT-2342855' → '2342855'. Si no coincide devuelve el original."""
    m = re.search(r"NIT-?(\w+)", raw_id, re.IGNORECASE)
    return m.group(1) if m else raw_id


def split_name(full_name: str) -> tuple[str, str]:
    """
    Divide un nombre completo guatemalteco en (nombres, apellidos).
    Heurística: los últimas 2 palabras son apellidos; el resto nombres.
    """
    parts = full_name.strip().split()
    if len(parts) <= 1:
        return full_name.strip(), ""
    if len(parts) == 2:
        return parts[0], parts[1]
    return " ".join(parts[:-2]), " ".join(parts[-2:])


def read_csv_file(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path: str, fieldnames: list[str], rows: list[dict]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  CSV  → {os.path.relpath(path, BASE)}  ({len(rows)} filas)")


def build_insert_chunks(table: str, columns: list[str],
                        rows: list[dict], chunk: int = 500) -> list[str]:
    """Genera sentencias INSERT INTO ... VALUES (...) en lotes de `chunk` filas."""
    if not rows:
        return []
    cols_str = ", ".join(columns)
    statements = []
    for i in range(0, len(rows), chunk):
        batch = rows[i:i + chunk]
        values = []
        for row in batch:
            vals = ", ".join(esc(row.get(c)) for c in columns)
            values.append(f"  ({vals})")
        stmt = f"INSERT INTO {table} ({cols_str}) VALUES\n" + ",\n".join(values) + ";"
        statements.append(stmt)
    return statements


def write_sql(path: str, header: str, sections: list[tuple[str, list[str]]]):
    """
    Escribe un archivo SQL.
    sections = [(comentario, [sentencias_sql]), ...]
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"-- {header}",
        f"-- Generado: {ts}",
        "",
        "USE ods_justicia;",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "",
    ]
    for comment, stmts in sections:
        if comment:
            lines.append(f"-- {'='*60}")
            lines.append(f"-- {comment}")
            lines.append(f"-- {'='*60}")
        lines.extend(stmts)
        lines.append("")
    lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  SQL  → {os.path.relpath(path, BASE)}")


# ══════════════════════════════════════════════════════════════════════════════
# FUENTE 1: OCDS (contratos/)
# Tablas: Tipo_Contrato, Institucion, Empresa, Contrato
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("FUENTE 1: contratos/ (CSVs OCDS)")
print("="*60)

# ── Leer archivos de tamaño manejable primero ─────────────────────────────────
contract_rows = read_csv_file(os.path.join(DIR_CONT, "contracts.csv"))
supplier_rows = read_csv_file(os.path.join(DIR_CONT, "awards_suppliers.csv"))
print(f"  contracts.csv:        {len(contract_rows):>6} filas")
print(f"  awards_suppliers.csv: {len(supplier_rows):>6} filas")

contracts_by_ocid: dict[str, list[dict]] = {}
for r in contract_rows:
    contracts_by_ocid.setdefault(r.get("main_ocid", ""), []).append(r)

suppliers_by_ocid: dict[str, list[dict]] = {}
for r in supplier_rows:
    suppliers_by_ocid.setdefault(r.get("main_ocid", ""), []).append(r)

valid_ocids = set(contracts_by_ocid.keys())

# ── Streaming: parties.csv → solo compradores (evita cargar 6 M filas en RAM) ─
print("  Leyendo parties.csv (streaming, solo buyers) ...", end="", flush=True)
buyer_parties: dict[str, dict] = {}
_pcnt = 0
with open(os.path.join(DIR_CONT, "parties.csv"), newline="", encoding="utf-8-sig") as _f:
    for _r in csv.DictReader(_f):
        _pcnt += 1
        if "buyer" in _r.get("roles", "").lower():
            _bid = _r.get("identifier_id", "").strip()
            if _bid:
                buyer_parties[_bid] = {
                    "address_streetAddress":  _r.get("address_streetAddress", ""),
                    "contactPoint_telephone": _r.get("contactPoint_telephone", ""),
                }
print(f" {_pcnt:,} filas → {len(buyer_parties):,} compradores únicos")

# ── Streaming: main.csv → Tipo_Contrato, Institucion, índice ligero por OCID ──
# Solo almacena (buyer_id, proc_method) para los OCIDs con contratos;
# evita mantener 1.5 M filas completas en memoria.
print("  Leyendo main.csv (streaming) ...", end="", flush=True)
tipo_contrato_map: dict[str, int] = {}
inst_map:   dict[str, int]  = {}
inst_data:  dict[str, dict] = {}
ocid_index: dict[str, tuple] = {}   # ocid → (buyer_id, proc_method)
tc_id   = 1
inst_id = 1
_mcnt   = 0
with open(os.path.join(DIR_CONT, "main.csv"), newline="", encoding="utf-8-sig") as _f:
    for _r in csv.DictReader(_f):
        _mcnt += 1
        _ocid = _r.get("ocid",  "").strip()
        _bid  = _r.get("buyer_id", "").strip()
        _bnm  = _r.get("buyer_name", "").strip()
        _meth = _r.get("tender_procurementMethod", "").strip()

        # Tipo_Contrato
        if _meth and _meth not in tipo_contrato_map:
            tipo_contrato_map[_meth] = tc_id
            tc_id += 1

        # Institucion (buyer único por buyer_id)
        if _bid and _bid not in inst_map:
            _party = buyer_parties.get(_bid.replace("GT-NIT-", ""), {})
            inst_map[_bid] = inst_id
            inst_data[_bid] = {
                "id":                   inst_id,
                "id_municipio":         1,   # placeholder Guatemala
                "id_tipo_institucion":  1,   # placeholder
                "id_nivel_institucion": 1,   # placeholder
                "nombre":               _bnm,
                "siglas":               _bid,
                "direccion":            _party.get("address_streetAddress", "") or None,
                "telefono":             _party.get("contactPoint_telephone", "") or None,
                "fecha_creacion":       None,
            }
            inst_id += 1

        # Índice ligero solo para OCIDs que tienen contratos
        if _ocid in valid_ocids and _ocid not in ocid_index:
            ocid_index[_ocid] = (_bid, _meth)

print(f" {_mcnt:,} filas")

tipo_contrato_rows = [{"id": v, "nombre": k}
                      for k, v in sorted(tipo_contrato_map.items(), key=lambda x: x[1])]
inst_rows = list(inst_data.values())

# ── Empresa (proveedores OCDS) ─────────────────────────────────────────────────
emp_map: dict[str, int] = {}    # nit_raw → id
emp_data: dict[str, dict] = {}
emp_id = 1
for r in supplier_rows:
    sid = r.get("id", "").strip()
    snm = r.get("name", "").strip()
    if sid and sid not in emp_map:
        nit = extract_nit(sid)
        emp_map[sid] = emp_id
        emp_data[sid] = {
            "id":                   emp_id,
            "id_perfil_persona_repr": None,
            "id_tipo_empresa":      1,   # Sociedad Anónima por defecto
            "ingresos":             None,
            "egresos":              None,
            "nit":                  nit,
            "direccion":            None,
            "telefono":             None,
            "estado":               "Activo",
            "observaciones":        snm,   # nombre guardado como observación
            "fecha_registro":       None,
        }
        emp_id += 1

emp_rows_ocds = list(emp_data.values())

# Estado_Contrato mapping (según insert_ods.sql: 1=Activo, 2=Finalizado, 3=Rescindido, 4=Suspendido)
ESTADO_MAP = {"active": 1, "terminated": 2, "cancelled": 3, "suspended": 4}
DEFAULT_ESTADO = 1

# ── Contrato ───────────────────────────────────────────────────────────────────
contrato_rows = []
for ocid, ocid_contracts in contracts_by_ocid.items():
    info = ocid_index.get(ocid)
    if not info:
        continue
    buyer_id, proc_method = info

    inst_ref = inst_map.get(buyer_id)
    tc_ref   = tipo_contrato_map.get(proc_method)
    if not inst_ref or not tc_ref:
        continue

    for con in ocid_contracts:
        fecha_inicio = parse_date(con.get("dateSigned", ""))
        if not fecha_inicio:
            continue

        fecha_fin   = parse_date(con.get("period_endDate", ""))
        sueldo_real = con.get("value_amount", "0") or "0"
        try:
            sueldo_real = float(sueldo_real)
        except ValueError:
            sueldo_real = 0.0

        status     = con.get("status", "active").lower().strip()
        estado_ref = ESTADO_MAP.get(status, DEFAULT_ESTADO)

        # Primer proveedor del proceso
        sups    = suppliers_by_ocid.get(ocid, [])
        emp_ref = None
        if sups:
            sup_raw = sups[0].get("id", "").strip()
            emp_ref = emp_map.get(sup_raw)

        contrato_rows.append({
            "id_trabajador":      1,         # PLACEHOLDER
            "id_empresa":         emp_ref,
            "id_institucion":     inst_ref,
            "id_cargo":           1,         # PLACEHOLDER
            "id_tipo_contrato":   tc_ref,
            "id_estado_contrato": estado_ref,
            "sueldo_real":        sueldo_real,
            "fecha_inicio":       fecha_inicio,
            "fecha_fin":          fecha_fin,
        })

print(f"\nTablas OCDS extraídas:")
print(f"  Tipo_Contrato  : {len(tipo_contrato_rows)}")
print(f"  Institucion    : {len(inst_rows)}")
print(f"  Empresa (OCDS) : {len(emp_rows_ocds)}")
print(f"  Contrato       : {len(contrato_rows)}")

# ── Escribir CSV (OCDS) ────────────────────────────────────────────────────────
write_csv(os.path.join(OUT_CSV, "Tipo_Contrato.csv"),
          ["id", "nombre"], tipo_contrato_rows)

write_csv(os.path.join(OUT_CSV, "Institucion.csv"),
          ["id", "id_municipio", "id_tipo_institucion", "id_nivel_institucion",
           "nombre", "siglas", "direccion", "telefono", "fecha_creacion"],
          inst_rows)

write_csv(os.path.join(OUT_CSV, "Empresa_OCDS.csv"),
          ["id", "id_perfil_persona_repr", "id_tipo_empresa", "ingresos", "egresos",
           "nit", "direccion", "telefono", "estado", "observaciones", "fecha_registro"],
          emp_rows_ocds)

write_csv(os.path.join(OUT_CSV, "Contrato.csv"),
          ["id_trabajador", "id_empresa", "id_institucion", "id_cargo",
           "id_tipo_contrato", "id_estado_contrato", "sueldo_real",
           "fecha_inicio", "fecha_fin"],
          contrato_rows)

# ── Escribir SQL (OCDS) ────────────────────────────────────────────────────────
sql_ocds_sections = [
    ("Tipo_Contrato — desde tender_procurementMethod (main.csv)",
     build_insert_chunks("Tipo_Contrato", ["id", "nombre"], tipo_contrato_rows)),

    ("Institucion (buyers) — desde buyer_id/buyer_name + parties.csv\n"
     "-- NOTA: id_municipio, id_tipo_institucion, id_nivel_institucion = 1 (placeholder).",
     build_insert_chunks(
         "Institucion",
         ["id", "id_municipio", "id_tipo_institucion", "id_nivel_institucion",
          "nombre", "siglas", "direccion", "telefono"],
         inst_rows)),

    ("Empresa (proveedores OCDS) — desde awards_suppliers.csv\n"
     "-- El nombre del proveedor se guarda en la columna 'observaciones'.",
     build_insert_chunks(
         "Empresa",
         ["id", "id_perfil_persona_repr", "id_tipo_empresa",
          "nit", "estado", "observaciones"],
         emp_rows_ocds)),

    ("Contrato — join de main + contracts + awards_suppliers\n"
     "-- id_trabajador=1 e id_cargo=1 son placeholders (OCDS no tiene datos de RRHH).",
     build_insert_chunks(
         "Contrato",
         ["id_trabajador", "id_empresa", "id_institucion", "id_cargo",
          "id_tipo_contrato", "id_estado_contrato", "sueldo_real",
          "fecha_inicio", "fecha_fin"],
         contrato_rows)),
]

write_sql(os.path.join(OUT_SQL, "01_insert_contratos_ocds.sql"),
          "INSERT OCDS → Tipo_Contrato, Institucion, Empresa, Contrato",
          sql_ocds_sections)


# ══════════════════════════════════════════════════════════════════════════════
# FUENTE 2: finanzas/registro_de_proveedores.xlsx
# Tablas: Perfil_Persona (individuos), Empresa (todos los proveedores)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("FUENTE 2: registro_de_proveedores.xlsx")
print("="*60)

# Mapeo CLASIFICACION → id_tipo_empresa
TIPO_EMP_MAP = {
    "Persona Individual":     2,   # Empresa Individual
    "Comerciante Individual": 2,
    "Sociedades":             1,   # Sociedad Anónima
    "Sociedades Civiles":     3,   # SRL
    "ONG":                    5,
    "Asociación":             5,
    "Fundación":              5,
    "Cooperativas":           4,
}
INDIVIDUALES = {"Persona Individual", "Comerciante Individual"}

wb = openpyxl.load_workbook(
    os.path.join(DIR_FIN, "registro_de_proveedores.xlsx"), read_only=True
)
ws = wb["listado_proveedores"]

perfil_rows   = []
empresa_p_rows = []

# IDs: Perfil_Persona parte de 1; Empresa parte de 10_001 (para no colisionar con OCDS)
pp_id  = 1
emp2_id = 10_001

seen_nit: set[str] = set()   # evitar duplicados por NIT

for row in ws.iter_rows(min_row=2, values_only=True):
    nit_raw, nombre_rs, clasificacion, estado_rgae, fecha_ins, _, _ = (
        row[0], row[1], row[2], row[3], row[4], row[5], row[6]
    )

    nit          = str(nit_raw).strip() if nit_raw else None
    nombre_rs    = str(nombre_rs).strip() if nombre_rs else ""
    clasificacion = str(clasificacion).strip() if clasificacion else "Persona Individual"
    estado_str   = str(estado_rgae).strip() if estado_rgae else None
    fecha_str    = parse_date(fecha_ins)
    tipo_emp     = TIPO_EMP_MAP.get(clasificacion, 2)

    if not nit or nit in seen_nit:
        continue
    seen_nit.add(nit)

    id_perfil_ref = None

    if clasificacion in INDIVIDUALES:
        nombres, apellidos = split_name(nombre_rs)
        perfil_rows.append({
            "id":                    pp_id,
            "id_municipio_vivienda": None,
            "id_sexo":               None,
            "id_estado_civil":       None,
            "id_grupo_etnico":       None,
            "dpi":                   None,
            "nit":                   nit,
            "nombres":               nombres,
            "apellidos":             apellidos,
            "fecha_nacimiento":      None,
            "direccion":             None,
            "telefono":              None,
        })
        id_perfil_ref = pp_id
        pp_id += 1

    empresa_p_rows.append({
        "id":                    emp2_id,
        "id_perfil_persona_repr": id_perfil_ref,
        "id_tipo_empresa":       tipo_emp,
        "ingresos":              None,
        "egresos":               None,
        "nit":                   nit,
        "direccion":             None,
        "telefono":              None,
        "estado":                estado_str,
        "observaciones":         nombre_rs,   # nombre guardado como observación
        "fecha_registro":        fecha_str,
    })
    emp2_id += 1

wb.close()

print(f"\nTablas proveedores extraídas:")
print(f"  Perfil_Persona : {len(perfil_rows)}")
print(f"  Empresa        : {len(empresa_p_rows)}")

# ── Escribir CSV (proveedores) ─────────────────────────────────────────────────
write_csv(os.path.join(OUT_CSV, "Perfil_Persona.csv"),
          ["id", "id_municipio_vivienda", "id_sexo", "id_estado_civil",
           "id_grupo_etnico", "dpi", "nit", "nombres", "apellidos",
           "fecha_nacimiento", "direccion", "telefono"],
          perfil_rows)

write_csv(os.path.join(OUT_CSV, "Empresa_Proveedores.csv"),
          ["id", "id_perfil_persona_repr", "id_tipo_empresa", "ingresos", "egresos",
           "nit", "direccion", "telefono", "estado", "observaciones", "fecha_registro"],
          empresa_p_rows)

# ── Escribir SQL (proveedores) ─────────────────────────────────────────────────
sql_prov_sections = [
    ("Perfil_Persona — personas individuales del RGAE\n"
     "-- Nombres/apellidos extraídos de NOMBRE_O_RAZON_SOCIAL (últimas 2 palabras = apellidos).",
     build_insert_chunks(
         "Perfil_Persona",
         ["id", "nit", "nombres", "apellidos"],
         perfil_rows)),

    ("Empresa — todos los proveedores del RGAE\n"
     "-- El nombre comercial/razón social se guarda en la columna 'observaciones'.\n"
     "-- Los IDs de Empresa parten de 10001 para no colisionar con los de OCDS.",
     build_insert_chunks(
         "Empresa",
         ["id", "id_perfil_persona_repr", "id_tipo_empresa",
          "nit", "estado", "observaciones", "fecha_registro"],
         empresa_p_rows)),
]

write_sql(os.path.join(OUT_SQL, "02_insert_proveedores.sql"),
          "INSERT RGAE → Perfil_Persona, Empresa",
          sql_prov_sections)


# ══════════════════════════════════════════════════════════════════════════════
# FUENTE 3: ine/procesos_judiciales.xlsx
# Tabla: Indicador_Externo
# Cuadros 4-10: detenidos por delito y departamento (2024)
# Cuadro  11:   tasa de incidencia por sexo y año (2020-2024)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("FUENTE 3: procesos_judiciales.xlsx (INE)")
print("="*60)

CUADROS_DELITO = {
    "Cuadro 4":  "Detenidos por homicidio",
    "Cuadro 5":  "Detenidos por lesiones",
    "Cuadro 6":  "Detenidos por robos",
    "Cuadro 7":  "Detenidos por otros robos y hurtos",
    "Cuadro 8":  "Detenidos por secuestros",
    "Cuadro 9":  "Detenidos por extorsiones",
    "Cuadro 10": "Detenidos por otras causas",
}

DEPARTAMENTOS_GT = {
    "Total País", "Total paφs", "Total país",
    # skip these
}

wb_ine = openpyxl.load_workbook(
    os.path.join(DIR_INE, "procesos_judiciales.xlsx"), read_only=True
)

indicador_rows = []
ind_id = 1

def clean_str(val) -> str:
    """Limpia el valor leído del xlsx (puede venir con encoding raro)."""
    if val is None:
        return ""
    return str(val).strip()


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── Cuadros 4-10: detenidos por departamento ──────────────────────────────────
for sheet_name, tipo_indicador in CUADROS_DELITO.items():
    ws_c = wb_ine[sheet_name]
    # Estructura: fila header ≈ row 7 ("Departamento1/", "Total", ...)
    # Fila 8: (None, "Ambos sexos", "Hombre", "Mujer", ...)
    # Filas 9+: (depto, total_ambos, total_h, total_m, enero_ambos, ...)
    data_rows = list(ws_c.iter_rows(values_only=True))

    # Encontrar la fila de inicio de datos (primera que tiene "Total" o nombre de depto en col 1)
    data_start = None
    for i, r in enumerate(data_rows):
        val0 = clean_str(r[0]).lower()
        if "total pa" in val0:
            data_start = i
            break

    if data_start is None:
        print(f"  [AVISO] No se encontró inicio de datos en {sheet_name}")
        continue

    for r in data_rows[data_start:]:
        depto = clean_str(r[0])
        if not depto:
            break
        if any(kw in depto.lower() for kw in ["fuente", "nota", "1/los"]):
            break
        if "total pa" in depto.lower():
            continue   # skipear totales país

        total_ambos = safe_float(r[1])
        total_h     = safe_float(r[2])
        total_m     = safe_float(r[3])

        for valor, unidad_sfx in [
            (total_ambos, "Ambos sexos"),
            (total_h,     "Hombres"),
            (total_m,     "Mujeres"),
        ]:
            if valor is None:
                continue
            indicador_rows.append({
                "id":             ind_id,
                "anio":           2024,
                "fuente":         "INE - Policía Nacional Civil",
                "tipo_indicador": tipo_indicador,
                "valor":          round(valor, 4),
                "unidad":         f"Detenidos ({unidad_sfx}) — {depto}",
            })
            ind_id += 1

# ── Cuadro 11: tasa de incidencia por año ─────────────────────────────────────
ws_11 = wb_ine["Cuadro 11"]
data_rows_11 = list(ws_11.iter_rows(values_only=True))

# Estructura: Año | Tasa Ambos sexos | Hombres | Mujeres
# Filas de datos empiezan cuando col 0 es un entero (año)
for r in data_rows_11:
    anio_val = r[0]
    if not isinstance(anio_val, (int, float)):
        continue
    anio       = int(anio_val)
    tasa_ambos = safe_float(r[1])
    tasa_h     = safe_float(r[2])
    tasa_m     = safe_float(r[3])

    for valor, sexo in [
        (tasa_ambos, "Ambos sexos"),
        (tasa_h,     "Hombres"),
        (tasa_m,     "Mujeres"),
    ]:
        if valor is None:
            continue
        indicador_rows.append({
            "id":             ind_id,
            "anio":           anio,
            "fuente":         "INE — Tasa incidencia delictiva",
            "tipo_indicador": "Tasa de incidencia delictiva",
            "valor":          round(valor, 4),
            "unidad":         f"Por 100,000 hab. ({sexo})",
        })
        ind_id += 1

wb_ine.close()

print(f"\nTabla INE extraída:")
print(f"  Indicador_Externo : {len(indicador_rows)}")

# ── Escribir CSV (INE) ─────────────────────────────────────────────────────────
write_csv(os.path.join(OUT_CSV, "Indicador_Externo.csv"),
          ["id", "anio", "fuente", "tipo_indicador", "valor", "unidad"],
          indicador_rows)

# ── Escribir SQL (INE) ─────────────────────────────────────────────────────────
sql_ine_sections = [
    ("Indicador_Externo — INE / Cuadros 4-10 (detenidos 2024) + Cuadro 11 (tasas 2020-2024).",
     build_insert_chunks(
         "Indicador_Externo",
         ["id", "anio", "fuente", "tipo_indicador", "valor", "unidad"],
         indicador_rows)),
]

write_sql(os.path.join(OUT_SQL, "03_insert_indicadores_ine.sql"),
          "INSERT INE → Indicador_Externo",
          sql_ine_sections)


# ══════════════════════════════════════════════════════════════════════════════
# RESUMEN FINAL
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("RESUMEN")
print("="*60)
print(f"  Tipo_Contrato      CSV  : {len(tipo_contrato_rows)} filas")
print(f"  Institucion        CSV  : {len(inst_rows)} filas")
print(f"  Empresa (OCDS)     CSV  : {len(emp_rows_ocds)} filas")
print(f"  Contrato           CSV  : {len(contrato_rows)} filas")
print(f"  Perfil_Persona     CSV  : {len(perfil_rows)} filas")
print(f"  Empresa (RGAE)     CSV  : {len(empresa_p_rows)} filas")
print(f"  Indicador_Externo  CSV  : {len(indicador_rows)} filas")
print(f"\nArchivos SQL generados en: output/sql/")
print(f"  01_insert_contratos_ocds.sql")
print(f"  02_insert_proveedores.sql")
print(f"  03_insert_indicadores_ine.sql")
print(f"\nEjecución completada. {datetime.now().strftime('%H:%M:%S')}")
